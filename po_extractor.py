"""
Purchase Order (PO) OCR Extraction Tool
Extracts data from PO PDFs and exports to Excel format.
"""

import os
import re
import glob
import pdfplumber
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment


# State codes mapping (first 2 digits of GSTIN)
STATE_CODES = {
    '01': 'Jammu & Kashmir', '02': 'Himachal Pradesh', '03': 'Punjab',
    '04': 'Chandigarh', '05': 'Uttarakhand', '06': 'Haryana',
    '07': 'Delhi', '08': 'Rajasthan', '09': 'Uttar Pradesh',
    '10': 'Bihar', '11': 'Sikkim', '12': 'Arunachal Pradesh',
    '13': 'Nagaland', '14': 'Manipur', '15': 'Mizoram',
    '16': 'Tripura', '17': 'Meghalaya', '18': 'Assam',
    '19': 'West Bengal', '20': 'Jharkhand', '21': 'Odisha',
    '22': 'Chhattisgarh', '23': 'Madhya Pradesh', '24': 'Gujarat',
    '26': 'Dadra & Nagar Haveli', '27': 'Maharashtra', '29': 'Karnataka',
    '30': 'Goa', '31': 'Lakshadweep', '32': 'Kerala',
    '33': 'Tamil Nadu', '34': 'Puducherry', '35': 'Andaman & Nicobar',
    '36': 'Telangana', '37': 'Andhra Pradesh'
}


def _format_date_mm_dd_yyyy(date_str):
    """Parse date string (dd/mm/yyyy or dd-mm-yyyy) and return mm/dd/yyyy."""
    if not date_str:
        return ''
    normalized = date_str.replace('.', '/').replace('-', '/')
    parts = normalized.split('/')
    if len(parts) != 3:
        return date_str
    try:
        # Indian PO format is typically dd/mm/yyyy
        day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
        d = datetime(year, month, day)
        return d.strftime('%m/%d/%Y')
    except (ValueError, TypeError):
        return date_str


def extract_po_data(pdf_path):
    """Extract purchase order data from a PDF file. Returns a list of rows (one per article line)."""
    def empty_row():
        return {
            'CHAINS': '', 'SITE CODE': '', 'STATE': '', 'VENDOR CODE': '', 'VENDOR NAME': '',
            'Sales Person': '', 'PO NO': '', 'PO DATE': '', 'DELIVERY DATE': '',
            'ARTICLE DESCRIPTION': '', 'TOTAL PCS': '', 'BASIC PRICE WITHOUT TAX': '',
            'TOTAL BASIC PO VALUE WITHOUT TAX': '', 'REMARKS': '', 'GRN AMOUNT': '',
            'Price/pcs': '', 'Actual Billing Price': '', 'Billing price of Reliance': '',
            'Price Difference': '', 'Remarks By SO': ''
        }

    # Article description: capture full text including weight in brackets e.g. SHAREAT FOOCHKA PANI PURI(1KG)
    # Lookahead (?=...\s+\d) ensures we only end at UOM when followed by quantity (not "KG" inside "(1KG)")
    article_regex = re.compile(
        r'(\d{13})\s+'  # EAN (13 digits)
        r'([\w\s\(\)\-/]+?)\s*'  # Article description (incl. weight like (1KG))
        r'(?=(?:EA|PC|KG|LT|MT)\s+\d)'  # End only when UOM is followed by quantity
        r'(?:EA|PC|KG|LT|MT)\s+'  # Unit of measure
        r'(\d+)\s+'  # Quantity
        r'\d+\s+'  # Free
        r'[\d.]+\s+'  # Basic Price (B.Price)
        r'[\d.]+\s+'  # Special Discount
        r'[\d.]+\s+'  # Schedule Value
        r'[\d.]+\s+'  # SGST
        r'[\d.]+\s+'  # CGST
        r'[\d.]+\s+'  # Cess
        r'([\d.]+)\s+'  # Landing Price (L.Price)
        r'[\d.]+\s+'  # MRP
        r'([\d,]+\.?\d*)',  # Total Value
        re.MULTILINE
    )

    rows = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ''
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + '\n'
            
            data = empty_row()
            
            # Extract Chain Name (Company name after "Ship To")
            chain_match = re.search(r'Ship To\s+([\w\s]+Ltd\.?)', full_text)
            if chain_match:
                data['CHAINS'] = chain_match.group(1).strip()
            
            # Extract Site Code / Address (location details)
            site_code = ''
            block_match = re.search(r'Ship To\s+(.*?)CIN:', full_text, re.DOTALL)
            if block_match:
                block = block_match.group(1)
                block = re.sub(r'Avenue Supermarts Ltd\.?', '', block)
                block = re.sub(r'PO\s*#\s*\d+', '', block)
                block = re.sub(r'PO\s*Date\s*[\d.]+', '', block)
                block = re.sub(r'Delivery\s*Dt?\s*[\d.]+', '', block)
                block = re.sub(r'\n', ' ', block)
                block = re.sub(r'\s+', ' ', block).strip()
                parts = []
                dmart_match = re.search(r'([A-Za-z\s]+(?:DMart|Dmart|DMART))', block)
                if dmart_match:
                    parts.append(dmart_match.group(1).strip())
                    block = block.replace(dmart_match.group(1), '')
                addr_parts = re.findall(r'([A-Za-z][A-Za-z\s,]+?)(?=\s+[A-Z][a-z]+\s+\d{6}|\s*$)', block)
                for part in addr_parts:
                    cleaned = part.strip(' ,')
                    if cleaned and cleaned not in parts:
                        parts.append(cleaned)
                city_pin = re.search(r'([A-Z][a-z]+)\s+(\d{6})', block)
                if city_pin:
                    parts.append(f"{city_pin.group(1)} - {city_pin.group(2)}")
                site_code = ', '.join(parts)
                site_code = re.sub(r',\s*,', ',', site_code)
                site_code = re.sub(r'\s+', ' ', site_code).strip(' ,')
            data['SITE CODE'] = site_code
            
            # Extract PO Number
            po_match = re.search(r'PO\s*#\s*(\d+)', full_text)
            if po_match:
                data['PO NO'] = po_match.group(1)
            
            # Extract PO Date (output format mm/dd/yyyy)
            po_date_match = re.search(r'PO\s*Date\s*(\d{2}[./]\d{2}[./]\d{4})', full_text)
            if po_date_match:
                data['PO DATE'] = _format_date_mm_dd_yyyy(po_date_match.group(1))
            
            # Extract Delivery Date (output format mm/dd/yyyy)
            delivery_match = re.search(r'Delivery\s*Dt?\s*(\d{2}[./]\d{2}[./]\d{4})', full_text)
            if delivery_match:
                data['DELIVERY DATE'] = _format_date_mm_dd_yyyy(delivery_match.group(1))
            
            # Extract Vendor Name
            vendor_match = re.search(r'Vendor\s+([\w\s]+?)(?=\s+GSTIN|\s+Phone|\s+FSSAI|Email)', full_text)
            if vendor_match:
                data['VENDOR NAME'] = vendor_match.group(1).strip()
            else:
                vendor_match2 = re.search(r'Phone\s+Vendor\s+([\w\s]+?)(?=\s+GSTIN|Email|\n)', full_text)
                if vendor_match2:
                    data['VENDOR NAME'] = vendor_match2.group(1).strip()
            
            # Extract State from GSTIN (store/ship-to location)
            gstin_match = re.search(r'GSTIN[:\s]*(\d{2}[A-Z0-9]+)', full_text)
            if gstin_match:
                state_code = gstin_match.group(1)[:2]
                data['STATE'] = STATE_CODES.get(state_code, '')
            
            # Vendor code: do not use GST number. Look for explicit "Vendor Code" in PDF if present.
            vendor_code_match = re.search(r'Vendor\s*Code\s*[:\s]*([A-Za-z0-9\-]+)', full_text, re.IGNORECASE)
            if vendor_code_match:
                data['VENDOR CODE'] = vendor_code_match.group(1).strip()
            
            # Weight continuation on next line e.g. "PURI(1KG)" or "PANIPURI(200G)" after "SHAREAT FOOCHKA PANI"
            weight_continuation_re = re.compile(
                r'\n\s*([A-Za-z]*\s*\(\d+(?:\.\d+)?\s*(?:KG|G|ML|LT)\))\s*\[?',
                re.IGNORECASE
            )
            # Extract ALL article line items (multiple articles per PO)
            article_matches = list(article_regex.finditer(full_text))
            if article_matches:
                for m in article_matches:
                    row = data.copy()
                    desc = m.group(2).strip()
                    # Append weight-in-brackets from next line if present (e.g. "PURI(1KG)" or "PANIPURI(200G)")
                    after_match = full_text[m.end():m.end() + 120]
                    cont = weight_continuation_re.search(after_match)
                    if cont:
                        desc = desc + ' ' + cont.group(1).strip()
                    
                    if 'shareat' not in desc.lower():
                        continue

                    row['ARTICLE DESCRIPTION'] = desc
                    row['TOTAL PCS'] = m.group(3)
                    row['BASIC PRICE WITHOUT TAX'] = m.group(4)
                    row['TOTAL BASIC PO VALUE WITHOUT TAX'] = m.group(5).replace(',', '')
                    if row['ARTICLE DESCRIPTION']:
                        row['ARTICLE DESCRIPTION'] = re.sub(r'\[HSN.*?\]', '', row['ARTICLE DESCRIPTION']).strip()
                        row['ARTICLE DESCRIPTION'] = re.sub(r'\s+', ' ', row['ARTICLE DESCRIPTION'])
                    rows.append(row)
            else:
                # Fallback: single-article extraction with simpler patterns
                desc_match = re.search(r'(\d{13})\s+([\w\s\(\)\-/]+?)(?:\s*\[HSN|\s+EA\s+|\s+PC\s+|\s+KG\s+)', full_text)
                if desc_match:
                    data['ARTICLE DESCRIPTION'] = desc_match.group(2).strip()
                qty_match = re.search(r'(?:EA|PC|KG|LT|MT)\s+(\d+)\s+\d+\s+([\d.]+)', full_text)
                if qty_match:
                    data['TOTAL PCS'] = qty_match.group(1)
                    data['BASIC PRICE WITHOUT TAX'] = qty_match.group(2)
                total_match = re.search(r'Total\s+(\d+)\s+([\d,]+\.?\d*)', full_text)
                if total_match:
                    if not data['TOTAL PCS']:
                        data['TOTAL PCS'] = total_match.group(1)
                    data['TOTAL BASIC PO VALUE WITHOUT TAX'] = total_match.group(2).replace(',', '')
                if data['ARTICLE DESCRIPTION']:
                    data['ARTICLE DESCRIPTION'] = re.sub(r'\[HSN.*?\]', '', data['ARTICLE DESCRIPTION']).strip()
                    data['ARTICLE DESCRIPTION'] = re.sub(r'\s+', ' ', data['ARTICLE DESCRIPTION'])
                
                if 'shareat' in data.get('ARTICLE DESCRIPTION', '').lower():
                    rows.append(data)
            
    except Exception as e:
        print(f"Error processing {pdf_path}: {str(e)}")
        rows = [empty_row()]
    
    return rows


def process_po_folder(input_folder, output_file=None):
    """Process all PO PDFs in a folder and export to Excel."""
    
    # Find all PDF files (use set to avoid duplicates)
    pdf_files = set()
    for pdf in glob.glob(os.path.join(input_folder, '*.pdf')):
        pdf_files.add(os.path.abspath(pdf))
    for pdf in glob.glob(os.path.join(input_folder, '**', '*.pdf'), recursive=True):
        pdf_files.add(os.path.abspath(pdf))
    
    pdf_files = list(pdf_files)
    
    if not pdf_files:
        print(f"No PDF files found in {input_folder}")
        return None
    
    print(f"Found {len(pdf_files)} PDF file(s) to process...")
    
    # Extract data from all PDFs (each PDF can yield multiple rows, one per article)
    all_data = []
    for pdf_path in pdf_files:
        print(f"Processing: {os.path.basename(pdf_path)}")
        rows = extract_po_data(pdf_path)
        all_data.extend(rows)
    
    # Create DataFrame
    df = pd.DataFrame(all_data)
    
    # Define output file if not specified
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(input_folder, f'PO_Extracted_{timestamp}.xlsx')
    
    # Export to Excel
    df.to_excel(output_file, index=False, sheet_name='PO Data')
    
    # Format the Excel file
    format_excel_output(output_file)
    
    print(f"\nExtraction complete! Output saved to: {output_file}")
    return output_file


def format_excel_output(excel_path):
    """Format the Excel output file for better readability."""
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Wrap text for long columns
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        wb.save(excel_path)
    except Exception as e:
        print(f"Warning: Could not format Excel file: {str(e)}")


def main():
    """Main function to run the PO extraction tool."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Extract PO data from PDF files to Excel')
    parser.add_argument('--input', '-i', 
                        default='sample data',
                        help='Input folder containing PO PDF files (default: sample data)')
    parser.add_argument('--output', '-o',
                        help='Output Excel file path (default: auto-generated in input folder)')
    
    args = parser.parse_args()
    
    # Get the script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Resolve input folder path
    if os.path.isabs(args.input):
        input_folder = args.input
    else:
        input_folder = os.path.join(script_dir, args.input)
    
    if not os.path.exists(input_folder):
        print(f"Error: Input folder not found: {input_folder}")
        return
    
    # Process the folder
    output_file = args.output
    if output_file and not os.path.isabs(output_file):
        output_file = os.path.join(script_dir, output_file)
    
    result = process_po_folder(input_folder, output_file)
    
    if result:
        print("\nExtracted data preview:")
        df = pd.read_excel(result)
        # Show key columns
        key_cols = ['PO NO', 'PO DATE', 'VENDOR NAME', 'ARTICLE DESCRIPTION', 'TOTAL PCS', 'TOTAL BASIC PO VALUE WITHOUT TAX']
        existing_cols = [col for col in key_cols if col in df.columns]
        print(df[existing_cols].to_string(index=False))


if __name__ == '__main__':
    main()
